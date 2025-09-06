from io import BytesIO
from typing import List, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.errors import AppException, ErrorCode
from app.models.entities import Taxonomy, TaxonomyEntry

REQUIRED_COLS = {"tag", "type", "reference"}


def _validate_and_parse_excel(contents: bytes, sheet_name: str):
    wb = load_workbook(BytesIO(contents), read_only=True)
    if sheet_name not in wb.sheetnames:
        raise AppException(
            ErrorCode.FILE_VALIDATION_ERROR,
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}",
            status_code=400,
        )

    ws = wb[sheet_name]

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = REQUIRED_COLS - set(headers)
    if missing:
        raise AppException(
            ErrorCode.FILE_VALIDATION_ERROR,
            f"Missing required columns: {', '.join(missing)}",
            status_code=400,
        )

    col_idx = {header: idx for idx, header in enumerate(headers)}

    for row in ws.iter_rows(min_row=2, values_only=True): 
        yield {
            "tag": row[col_idx["tag"]],
            "type": row[col_idx["type"]],
            "reference": row[col_idx["reference"]],
        }


def upload_taxonomy_service(db: Session, file_contents: bytes, filename: str, sheet_name: str, taxonomy: Optional[str] = None, description: Optional[str] = None) -> int:
    taxonomy_obj = Taxonomy(name=sheet_name, taxonomy=taxonomy, description=description, source_file=filename)
    db.add(taxonomy_obj)
    db.commit()
    db.refresh(taxonomy_obj)

    for row in _validate_and_parse_excel(file_contents, sheet_name):
        entry = TaxonomyEntry(
            taxonomy_id=taxonomy_obj.id,
            tag=row["tag"],
            datatype=row["type"],
            reference=row["reference"],
        )
        db.add(entry)

    db.commit()
    return taxonomy_obj.id


# to get the taxonomy by symbol like (brsr, sasb)
def get_taxonomy_by_taxonomy_name(db: Session, taxonomy: str) -> Taxonomy:
    return db.query(Taxonomy).filter(Taxonomy.taxonomy == taxonomy).first()


def get_taxonomy_service(db: Session, taxonomy_id: int) -> Taxonomy:
    taxonomy = db.query(Taxonomy).filter(Taxonomy.id == taxonomy_id).first()
    if not taxonomy:
        raise AppException(ErrorCode.NOT_FOUND, "Taxonomy not found", status_code=404)
    return taxonomy


def delete_taxonomy_service(db: Session, taxonomy_id: int) -> None:
    taxonomy = db.query(Taxonomy).filter(Taxonomy.id == taxonomy_id).first()
    if not taxonomy:
        raise AppException(ErrorCode.NOT_FOUND, "Taxonomy not found", status_code=404)
    db.delete(taxonomy)
    db.commit()


def get_entries_service(db: Session, taxonomy_id: int) -> List[TaxonomyEntry]:
    return db.query(TaxonomyEntry).filter(TaxonomyEntry.taxonomy_id == taxonomy_id).all()


def add_entry_service(db: Session, taxonomy_id: int, tag: str, datatype: str, reference: str) -> None:
    db.add(TaxonomyEntry(taxonomy_id=taxonomy_id, tag=tag, datatype=datatype, reference=reference))
    db.commit()


def update_entry_service(db: Session, entry_id: int, tag: str, datatype: str, reference: str) -> None:
    entry = db.query(TaxonomyEntry).filter(TaxonomyEntry.id == entry_id).first()
    if not entry:
        raise AppException(ErrorCode.NOT_FOUND, "Entry not found", status_code=404)
    entry.tag = tag
    entry.datatype = datatype
    entry.reference = reference
    db.commit()


def delete_entry_service(db: Session, entry_id: int) -> None:
    entry = db.query(TaxonomyEntry).filter(TaxonomyEntry.id == entry_id).first()
    if not entry:
        raise AppException(ErrorCode.NOT_FOUND, "Entry not found", status_code=404)
    db.delete(entry)
    db.commit()



