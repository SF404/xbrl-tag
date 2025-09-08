from io import BytesIO
from typing import Iterable
from openpyxl import load_workbook
from app.core.errors import AppException, ErrorCode

REQUIRED_COLS = {"tag", "type", "reference"}

def validate_and_parse_excel(contents: bytes, sheet_name: str) -> Iterable[dict]:
    wb = load_workbook(BytesIO(contents), read_only=True)
    if sheet_name not in wb.sheetnames:
        raise AppException(
            ErrorCode.FILE_VALIDATION_ERROR,
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}",
            status_code=400,
        )
    ws = wb[sheet_name]
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = REQUIRED_COLS - set(headers)
    if missing:
        raise AppException(
            ErrorCode.FILE_VALIDATION_ERROR,
            f"Missing required columns: {', '.join(missing)}",
            status_code=400,
        )
    col_idx = {header: idx for idx, header in enumerate(headers)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield {
            "tag": row[col_idx["tag"]],
            "type": row[col_idx["type"]],
            "reference": row[col_idx["reference"]],
        }